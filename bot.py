# -*- coding: utf-8 -*-
import os
import io
import logging
import re
import csv
import time
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, Bot
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, ContextTypes, filters
)
import openpyxl
from scrapers import (
    youtube_scraper, 
    tiktok_post_details_scraper, 
    tiktok_channel_posts_scraper, 
    dailymotion_scraper, 
    okru_scraper
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_domain_fast(url):
    try:
        if not url.startswith(('http://', 'https://')):
            url = 'http://' + url
        parsed = urlparse(url)
        domain = parsed.netloc
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain
    except:
        return ""

def extract_domains_bulk(urls):
    results = []
    def extract_single(url):
        return {"URL": url, "Domain": extract_domain_fast(url)}
    
    with ThreadPoolExecutor(max_workers=50) as executor:
        future_to_url = {executor.submit(extract_single, url): url for url in urls}
        for future in as_completed(future_to_url):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                url = future_to_url[future]
                results.append({"URL": url, "Domain": ""})
                logger.error(f"Error extracting domain from {url}: {e}")
    return results

PLATFORMS = {
    "YouTube": youtube_scraper,
    "TikTok": None,
    "Dailymotion": dailymotion_scraper,
    "Ok.ru": okru_scraper,
    "Domain Extractor": None,
}

TEMPLATE_FILES = {
    "YouTube": "templates/YouTube-Template.xlsx",
    "TikTok": "templates/UGC-Template.xlsx",
    "Dailymotion": "templates/UGC-Template.xlsx",
    "Ok.ru": "templates/UGC-Template.xlsx",
}

SELECT_PLATFORM, SELECT_TIKTOK_MODE, GET_INPUT = range(3)

URL_PATTERNS = {
    "YouTube": re.compile(r"^(https?:\/\/)?(www\.)?((youtube\.com\/watch\?v=)|youtube\.com\/shorts\/|youtu\.be\/)[a-zA-Z0-9_-]{11}($|&|/|\?)"),
    "TikTok": re.compile(r"^(https?:\/\/)?(www\.)?tiktok\.com\/@[\w._-]+\/video\/\d+"),
    "TikTok_Profile": re.compile(r"^(https?:\/\/)?(www\.)?tiktok\.com\/@[\w._-]+$|^@?[\w._-]+$"),
    "Dailymotion": re.compile(r"^(https?:\/\/)?(www\.)?dailymotion\.com\/video\/[a-zA-Z0-9]+$"),
    "Ok.ru": re.compile(r"^(https?:\/\/)?(www\.)?ok\.ru\/video\/\d+$"),
    "Domain Extractor": re.compile(r".*"),
}

fields_mapping = {
    "YouTube": {
        "source_url": "A", "title": "E", "videoId": "F", "views": "G", "duration": "H",
        "channelId": "I", "channel_name": "J", "channel_subs": "K", "likes": "L", 
        "publish_date": "M", "channel_username": "T",
    },
    "TikTok": {
        "source_url": "A", "title": "B", "views": "C", "duration": "D", "likes": "F",
        "comments": "G", "upload_date": "H", "profile_url": "L", "author_name": "M",
        "subscribers": "N", "channel_username": "V",
    },
    "Dailymotion": {
        "source_url": "A", "title": "B", "views": "C", "duration": "D", "likes": "F",
        "upload_date": "H", "channel_url": "L", "channel_name": "M", "subscribers": "N",
        "channel_username": "V",
    },
    "Ok.ru": {
        "source_url": "A", "title": "B", "views": "C", "duration": "D", "likes": "F",
        "upload_date": "H", "channel_url": "L", "channel_name": "M", "subscribers": "N",
        "channel_username": "V",
    }
}

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("🎥 YouTube", callback_data="YouTube")],
        [InlineKeyboardButton("🎵 TikTok", callback_data="TikTok")],
        [InlineKeyboardButton("📺 Dailymotion", callback_data="Dailymotion")],
        [InlineKeyboardButton("🎬 Ok.ru", callback_data="Ok.ru")],
        [InlineKeyboardButton("🌐 Domain Extractor", callback_data="Domain Extractor")]
    ]
    await update.message.reply_text(
        "🚀 *Platform Scraper Bot*\n\nSelect a platform to scrape:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )
    return SELECT_PLATFORM

async def platform_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    platform = query.data
    context.user_data['platform'] = platform
    
    if platform == "TikTok":
        keyboard = [
            [InlineKeyboardButton("📹 Post Details Scraper", callback_data="tiktok_post_details")],
            [InlineKeyboardButton("📺 Channel Post Extractor", callback_data="tiktok_channel_posts")],
            [InlineKeyboardButton("⬅️ Back to Platforms", callback_data="back_to_platforms")]
        ]
        await query.edit_message_text(
            text="🎵 *TikTok Scraper Options*\n\n"
                 "Choose scraping mode:\n\n"
                 "📹 *Post Details Scraper*\n"
                 "Extract data from specific video URLs\n\n"
                 "📺 *Channel Post Extractor*\n"
                 "Extract all videos from profile(s)",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )
        return SELECT_TIKTOK_MODE
    
    await query.edit_message_text(
        f"✅ You selected *{platform}*.\n\n"
        f"📝 Send me {platform} URLs line-by-line or upload an Excel file with URLs.",
        parse_mode='Markdown'
    )
    return GET_INPUT

async def tiktok_mode_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    mode = query.data
    
    if mode == "back_to_platforms":
        keyboard = [
            [InlineKeyboardButton("🎥 YouTube", callback_data="YouTube")],
            [InlineKeyboardButton("🎵 TikTok", callback_data="TikTok")],
            [InlineKeyboardButton("📺 Dailymotion", callback_data="Dailymotion")],
            [InlineKeyboardButton("🎬 Ok.ru", callback_data="Ok.ru")],
            [InlineKeyboardButton("🌐 Domain Extractor", callback_data="Domain Extractor")]
        ]
        await query.edit_message_text(
            "🚀 *Platform Scraper Bot*\n\nSelect a platform to scrape:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='Markdown'
        )
        return SELECT_PLATFORM
    elif mode == "tiktok_post_details":
        context.user_data['tiktok_mode'] = 'post_details'
        await query.edit_message_text(
            text="📹 *TikTok Post Details Scraper*\n\n"
                 "Send me TikTok video URLs or upload an Excel file with video URLs\n\n"
                 "Example URLs:\n"
                 "• https://www.tiktok.com/@user/video/123456789\n"
                 "• Multiple URLs separated by newlines",
            parse_mode='Markdown'
        )
    elif mode == "tiktok_channel_posts":
        context.user_data['tiktok_mode'] = 'channel_posts'
        await query.edit_message_text(
            text="📺 *TikTok Channel Post Extractor*\n\n"
                 "Send me TikTok profile URLs/usernames or upload an Excel file\n\n"
                 "Supported formats:\n"
                 "• https://www.tiktok.com/@username\n"
                 "• @username\n"
                 "• username\n"
                 "• Multiple profiles separated by newlines",
            parse_mode='Markdown'
        )
    return GET_INPUT

def extract_urls_from_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    urls = []
    url_col_idx = None
    for cell in ws[1]:
        if cell.value and "url" in str(cell.value).lower():
            url_col_idx = cell.column - 1
            break
    if url_col_idx is None:
        return []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[url_col_idx]
        if val:
            urls.append(str(val))
    return urls

async def input_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    platform = context.user_data.get('platform')
    tiktok_mode = context.user_data.get('tiktok_mode')
    urls = []
    
    if update.message.document:
        doc = update.message.document
        if not doc.file_name.lower().endswith(('.xls', '.xlsx')):
            await update.message.reply_text("❌ Please upload an Excel file (.xls or .xlsx).")
            return GET_INPUT
        file_obj = await doc.get_file()
        file_bytes = await file_obj.download_as_bytearray()
        urls = extract_urls_from_excel(file_bytes)
        if not urls:
            await update.message.reply_text("❌ Could not extract URLs from Excel file.")
            return GET_INPUT
    else:
        urls = [url.strip() for url in update.message.text.strip().splitlines() if url.strip()]

    if not urls:
        await update.message.reply_text("❌ No URLs provided!")
        return GET_INPUT

    if platform == "TikTok" and tiktok_mode == "channel_posts":
        pattern = URL_PATTERNS.get("TikTok_Profile")
        invalid_urls = [u for u in urls if not pattern.match(u)]
        if invalid_urls:
            await update.message.reply_text(
                f"❌ Invalid profile URLs/usernames:\n" + "\n".join(invalid_urls[:5]) + 
                f"\n{'...' if len(invalid_urls) > 5 else ''}\n\nPlease send valid TikTok profiles."
            )
            return GET_INPUT
    elif platform != "Domain Extractor" and platform != "TikTok":
        pattern = URL_PATTERNS.get(platform)
        invalid_urls = [u for u in urls if not pattern.match(u)]
        if invalid_urls:
            await update.message.reply_text(
                f"❌ Invalid URLs:\n" + "\n".join(invalid_urls[:5]) + 
                f"\n{'...' if len(invalid_urls) > 5 else ''}\n\nPlease send valid URLs."
            )
            return GET_INPUT
    elif platform == "TikTok" and tiktok_mode == "post_details":
        pattern = URL_PATTERNS.get("TikTok")
        invalid_urls = [u for u in urls if not pattern.match(u)]
        if invalid_urls:
            await update.message.reply_text(
                f"❌ Invalid TikTok video URLs:\n" + "\n".join(invalid_urls[:5]) + 
                f"\n{'...' if len(invalid_urls) > 5 else ''}\n\nPlease send valid video URLs."
            )
            return GET_INPUT

    processing_msg = await update.message.reply_text(
        f"🔄 Processing {len(urls)} URLs for {platform}{'(' + tiktok_mode.replace('_', ' ').title() + ')' if tiktok_mode else ''}..."
    )

    if platform == "Domain Extractor":
        try:
            start_time = time.time()
            rows = extract_domains_bulk(urls)
            end_time = time.time()
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=["URL", "Domain"])
            writer.writeheader()
            writer.writerows(rows)
            output.seek(0)
            await processing_msg.delete()
            await update.message.reply_document(
                document=io.BytesIO(output.getvalue().encode()),
                filename="domains.csv",
                caption=f"✅ Domain extraction completed!\n\n"
                       f"📊 Processed: {len(urls)} URLs\n"
                       f"⏱️ Time taken: {end_time - start_time:.2f} seconds"
            )
            return ConversationHandler.END
        except Exception as e:
            logger.exception(f"Error in domain extraction: {e}")
            await processing_msg.edit_text(f"❌ Error during domain extraction: {e}")
            return ConversationHandler.END

    try:
        if platform == "TikTok":
            if tiktok_mode == "post_details":
                results = await tiktok_post_details_scraper(urls)
            elif tiktok_mode == "channel_posts":
                results = await tiktok_channel_posts_scraper(urls)
            else:
                await processing_msg.edit_text("❌ Invalid TikTok mode!")
                return ConversationHandler.END
        else:
            scraper_func = PLATFORMS.get(platform)
            if not scraper_func:
                await processing_msg.edit_text(f"❌ No scraper found for {platform}")
                return ConversationHandler.END
            results = await scraper_func(urls)
    except Exception as e:
        logger.exception(f"Error scraping {platform}: {e}")
        await processing_msg.edit_text(f"❌ Error during scraping: {e}")
        return ConversationHandler.END

    if not results:
        await processing_msg.edit_text("❌ No data scraped.")
        return ConversationHandler.END

    template_path = TEMPLATE_FILES.get(platform)
    if not template_path or not os.path.isfile(template_path):
        await processing_msg.edit_text("❌ Template file missing for this platform.")
        return ConversationHandler.END

    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        mapping = fields_mapping.get(platform)
        if not mapping:
            await processing_msg.edit_text("❌ No field mapping found for this platform.")
            return ConversationHandler.END

        start_row = 2
        for i, row_data in enumerate(results, start=start_row):
            for field, col in mapping.items():
                val = row_data.get(field, "N/A")
                ws[f"{col}{i}"].value = val

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_name = f"{platform.replace(' ', '_')}{'_' + tiktok_mode if tiktok_mode else ''}_scraped_output.xlsx"
        
        await processing_msg.delete()
        await update.message.reply_document(
            document=output, 
            filename=file_name,
            caption=f"✅ Scraping completed!\n\n"
                   f"📊 Total results: {len(results)}\n"
                   f"📅 Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                   f"Use /start to scrape again!"
        )
    except Exception as e:
        logger.exception(f"Error creating output file: {e}")
        await processing_msg.edit_text(f"❌ Error creating output file: {e}")
        return ConversationHandler.END

    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Cancelled. Use /start to begin again.")
    return ConversationHandler.END

def main():
    TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
    if not TOKEN:
        logger.error("Missing TELEGRAM_BOT_TOKEN environment variable")
        return

    app = ApplicationBuilder().token(TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            SELECT_PLATFORM: [CallbackQueryHandler(platform_selected)],
            SELECT_TIKTOK_MODE: [CallbackQueryHandler(tiktok_mode_selected)],
            GET_INPUT: [MessageHandler((filters.TEXT | filters.Document.ALL) & ~filters.COMMAND, input_received)]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True
    )
    app.add_handler(conv)
    logger.info("🤖 Bot is running...")
    app.run_polling(allowed_updates=None)

if __name__ == "__main__":
    main()
